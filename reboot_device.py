import os
import re
import subprocess
import time

import openpyxl
from serial import Serial

from logger import logger
from util import get_modem_port, get_device_connect, get_strf_time, network_ping_success


class RebootDevice:

    def __init__(self):
        self.is_double_band = self._device_type()
        self.all_double_band = ['G3', 'U3X','E1']
        self.modem_port = self._get_modem_port()
        self.logs_path = self._make_dir()
        self.folder_size = 0

    def creart_excel(self):
        try:
            excel_path = os.getcwd() + '\\' + 'detail_' + time.strftime('%Y-%m-%d %H-%M-%S',
                                                                        time.localtime(time.time())) + '.xlsx'
            logger.info(f'creat excel save path {excel_path}')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['开始时间', '结束时间', '总共耗时', '云卡IMSI', '种子卡IMSI', '种子卡注册网络'])
            wb.save(excel_path)
            return excel_path
        except Exception as e:
            logger.error(e)

    def _make_dir(self):
        log_path = os.path.dirname(os.path.abspath(__file__)) + "\\" + "logs"
        if not os.path.isdir(log_path):
            os.makedirs(log_path)
            logger.info(f'sdcard/logs folder is not exists ,creat it {log_path}')
        else:
            logger.info(f'sdcard/logs Folder already exists {log_path}')
        return log_path

    def get_folder_size(self):

        try:
            temp_folder_size = subprocess.getstatusoutput("adb shell du -sh sdcard/logs")[1]
            if temp_folder_size.find("error") != -1:
                logger.info("error: no devices/emulators found")
            else:
                self.folder_size = temp_folder_size.split('\t')[0][:-1]
            return self.folder_size
        except Exception as e:
            logger.error(f'Failed to get sdcard/logs {e}')

    def get_pull_logs(self,log_path):
        try:
            pull_path = log_path + " " + self.logs_path
            t = subprocess.getoutput(pull_path)
            if t.find("error") != -1:
                logger.info(f'adb pull sdcard/logs fail {t},do not del sdcard/logs file')
            else:
                logger.info(f'adb pull sdcard/logs successful{t}')
                self._del_dir()
        except Exception as e:
            print("")

    def _del_dir(self):
        del_path_1 = 'adb shell rm -rf sdcard/logs/diag_logs/*'
        del_path_2 = 'adb shell rm -rf sdcard/logs/ap_logs/*'
        d = subprocess.getoutput(del_path_1)
        c = subprocess.getoutput(del_path_2)
        if d.find('error') != -1:
            logger.info(f'adb del sdcard/logs fail {d}')
        else:
            logger.info(f'adb del sdcard/logs successful')

    def sendAtResult(self, at_cmd):
        port = self.modem_port
        try:
            ser = Serial(port, 9600, timeout=1)
            cmd = ''.join([at_cmd, '\r\n']).encode()
            ser.write(cmd)
            result = ser.readall().decode().replace('\r\n', ' ').replace('\t', ' ').replace('\r', ' ')

            ser.close()
            # print(result)
            logger.info(f'execute at commont result {result}')
            return result
        except Exception as e:
            logger.error(f'send at comm fail {e}')

    def _device_type(self):
        while True:
            if get_device_connect():
                try:
                    device_type = subprocess.check_output(
                        'adb shell getprop ucloud.oem.conf.device_name').decode().strip()
                    logger.info(f'find device type {device_type}')
                    if device_type == None:
                        logger.info('can not find out device type')
                        raise ()
                    else:
                        return device_type

                except Exception as e:
                    logger.error(f"getdevices type fail,try again!{e}")

    def _get_modem_port(self):
        while True:
            if get_device_connect():
                try:
                    modem_port = get_modem_port()
                    logger.info(f'find modem port {modem_port}')
                    return modem_port
                except Exception as e:
                    logger.error("An error occurred when obtaining the modem port. Try again")

    def open_logs_config(self):
        obj = subprocess.Popen(["adb", "shell"], shell=True, stdin=subprocess.PIPE, stderr=subprocess.PIPE,
                               stdout=subprocess.PIPE)
        obj.stdin.write(
            "echo localOpenAp=1'\n'localOpenBp=1'\n'localQxLogCfg=0'\n'localLogSize=1600'\n'persistSize=512'\n'apLogsConfig=main,system,radio,dmesg,event,tcp>sdcard/logs.cfg\n".encode(
                "utf-8"))
        obj.stdin.write('exit\n'.encode("utf-8"))
        inf,err = obj.communicate()
        if err.decode() != '':
            logger.error(f'Failed to create logs configuration file{err.decode("utf-8")}')
        else:
            logger.info(f'creat logs configuration successful!')

    def reboot(self, n):
        save_path = self.creart_excel()
        if self.is_double_band not in self.all_double_band:
            self.open_logs_config()

        temp_row = 2
        for i in range(n):
            temp_data = []
            while True:
                time.sleep(2)
                n = 1
                logger.info("******************************start******************************")
                if get_device_connect():
                    result_subprocess = subprocess.getstatusoutput("adb reboot")[0]
                    print("******************************start****************************")
                    if result_subprocess == 0:
                        break
                    else:
                        logger.info(f'Failed to run the reboot command. Procedure {result_subprocess}')
                if n % 10 == 0:
                    logger.info("can not find devices! check the device is connected")
                    print('can not find devices! check the device is connected')
                n = n + 1

            start_time = time.time()
            start_format_time = get_strf_time()
            print(f"第{i+1}次重启 开始时间：{start_format_time}")

            logger.info(f'The system reboot time:{i + 1}')
            logger.info(f'start adb reboot time {start_format_time}')
            time.sleep(20)
            network_ping_success()
            print("云卡网络ping通")

            end_time = time.time()
            end_format_time = get_strf_time()
            print(f"第{i+1}重启 结束时间：{end_format_time}")

            total_time = str(round((end_time - start_time), 2)) + 's'
            print(f"总共耗时：{total_time}")
            print("******************************end******************************")
            logger.info(f'vsim ping success time:{i + 1},time {end_format_time}')
            logger.info(f'total user time{total_time}')
            try:
                t_vsim_imsi = self.sendAtResult('at+cimi')
                vsim_imsi = re.search(r'\d\d\d\d\d\d\d\d\d\d\d\d\d\d\d', t_vsim_imsi, re.S).group(0)
            except Exception as e:
                logger.info(e)
                vsim_imsi = None

            if self.is_double_band in self.all_double_band:
                self.sendAtResult("at^setat2ap=1")
                time.sleep(0.5)

                try:
                    t_softsim_imsi = self.sendAtResult("at2ap+cimi")
                    softsim_imsi = re.search(r'\d\d\d\d\d\d\d\d\d\d\d\d\d\d\d', t_softsim_imsi, re.S).group(0)
                except Exception as e:
                    softsim_imsi = None

                try:
                    time.sleep(1.5)
                    tem_softsim_cops = self.sendAtResult("at2ap+cops?")
                    softsim_cops = re.search(r'.*\d,(".*",\d+)', tem_softsim_cops, re.S).group(1)
                except Exception as e:
                    softsim_cops = None

            else:
                self.sendAtResult("at$qcsimapp=1")
                time.sleep(0.5)

                try:
                    t_softsim_imsi = self.sendAtResult("at+cimi")
                    softsim_imsi = re.search(r'\d\d\d\d\d\d\d\d\d\d\d\d\d\d\d', t_softsim_imsi, re.S).group(0)
                except Exception as e:
                    softsim_imsi = None

                try:
                    tem_softsim_cops = self.sendAtResult("at+cops?")
                    softsim_cops = re.search(r'.*\d,(".*",\d+)', tem_softsim_cops, re.S).group(1)
                except Exception as e:
                    softsim_cops = None

            temp_data = [start_format_time, end_format_time, total_time, vsim_imsi, softsim_imsi,
                         softsim_cops]
            logger.info(temp_data)

            try:
                wb = openpyxl.load_workbook(save_path)
                ws = wb.active
                for k in range(len(temp_data)):
                    ws.cell(row=temp_row, column=k + 1, value=temp_data[k])
                wb.save(save_path)
                temp_row = temp_row + 1
            except Exception as e:
                logger.error(f'open or save excel fail{e}')
            if self.is_double_band not in self.all_double_band:
                folder_size = self.get_folder_size()
                if float(folder_size) > 200:
                    self.get_pull_logs("adb pull sdcard/logs")
        self.get_pull_logs("adb pull sdcard/logs")
        self.get_pull_logs("adb pull sdcard/uaflogs")
        subprocess.Popen("pause",shell=True)

if __name__ == '__main__':
    s = input("重启次数：")
    test = RebootDevice()
    # test.creart_excel()
    test.reboot(int(s))
    # print(test.device_type())
    # test.sendAtResult('at+cimi')
    # print(type(test.get_folder_size()))
    # test.open_logs_config()
